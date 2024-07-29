from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime
from tkcalendar import Calendar
from dni_database import DniDatabase
from fpdf import FPDF
import openpyxl
import subprocess

class FrmReporteDNI:
    def __init__(self, root, restart_callback):
        self.root = root
        self.root.title("Reporte de asistencia diario y por fecha")
        self.root.configure(bg="#f0f0f0")
        self.restart_callback = restart_callback

        self.db = DniDatabase()

        # Configurar estilo
        style = ttk.Style()
        style.theme_use("clam")
        
        # Estilo para Treeview
        style.configure("Treeview", 
                        background="#d9d9d9",
                        foreground="black",
                        rowheight=25,
                        fieldbackground="#d9d9d9")
        style.map('Treeview', 
                  background=[('selected', '#347083')])

        # Estilo para los botones
        style.configure('TButton', 
                        background='#5f9ea0', 
                        foreground='white', 
                        font=('Arial', 10, 'bold'))
        style.map('TButton', 
                  background=[('active', '#4f8e90')])

        self.btnBuscar = ttk.Button(root, text="Buscar por Fecha", command=self.abrir_calendario)
        self.btnBuscar.grid(row=0, column=0, columnspan=2, pady=10, padx=10)

        self.lblRegistrosHoy = Label(root, text="Registros del Día:", bg="#f0f0f0", font=('Arial', 12, 'bold'))
        self.lblRegistrosHoy.grid(row=1, column=0, columnspan=2, pady=5)

        self.tree = ttk.Treeview(root, columns=("DNI", "Apellido Paterno", "Apellido Materno", "Nombres", "Fecha y Hora de Ingreso", "Fecha y Hora de Salida"), show='headings')
        self.tree.heading("DNI", text="DNI")
        self.tree.heading("Apellido Paterno", text="Apellido Paterno")
        self.tree.heading("Apellido Materno", text="Apellido Materno")
        self.tree.heading("Nombres", text="Nombres")
        self.tree.heading("Fecha y Hora de Ingreso", text="Fecha y Hora de Ingreso")
        self.tree.heading("Fecha y Hora de Salida", text="Fecha y Hora de Salida")
        self.tree.grid(row=2, column=0, columnspan=2, pady=10, padx=10)

        self.btnPDF = ttk.Button(root, text="Convertir a PDF", command=self.convertir_a_pdf)
        self.btnPDF.grid(row=3, column=0, pady=10, padx=10, sticky='ew')

        self.btnExcel = ttk.Button(root, text="Convertir a Excel", command=self.convertir_a_excel)
        self.btnExcel.grid(row=3, column=1, pady=10, padx=10, sticky='ew')

        self.btnBackToLogin = ttk.Button(root, text="Cerrar sesión", command=self.volver_al_login)
        self.btnBackToLogin.grid(row=0, column=1, columnspan=1, pady=10, padx=10, sticky='e')
        
        # Asociar la tecla "Esc" al botón de cerrar sesión
        self.root.bind('<Escape>', self.on_escape_key)

        self.btnRegistroAsistencia = ttk.Button(root, text="Reporte Mensual", command=self.ir_a_registro_asistencia)
        self.btnRegistroAsistencia.grid(row=4, column=0, columnspan=2, pady=10, padx=10, sticky='ew')

        self.load_today_records()
        
    def on_escape_key(self, event):
        self.volver_al_login()

    def volver_al_login(self):
        self.root.destroy()
        self.restart_callback()

    def abrir_calendario(self):
        self.cal_window = Toplevel(self.root)
        self.cal_window.title("Seleccione el Intervalo de Fechas")
        self.cal_window.geometry("400x510")
        self.cal_window.configure(bg="#f0f0f0")

        self.lblFechaInicio = Label(self.cal_window, text="Fecha de Inicio:", bg="#f0f0f0", font=('Arial', 10, 'bold'))
        self.lblFechaInicio.pack(pady=5)
        
        self.cal_inicio = Calendar(self.cal_window, selectmode='day', date_pattern='yyyy-mm-dd', mindate=datetime(2020, 1, 1), maxdate=datetime(2025, 12, 31))
        self.cal_inicio.pack(pady=5)

        self.lblFechaFin = Label(self.cal_window, text="Fecha de Fin:", bg="#f0f0f0", font=('Arial', 10, 'bold'))
        self.lblFechaFin.pack(pady=5)

        self.cal_fin = Calendar(self.cal_window, selectmode='day', date_pattern='yyyy-mm-dd', mindate=datetime(2020, 1, 1), maxdate=datetime(2025, 12, 31))
        self.cal_fin.pack(pady=5)

        self.btnSeleccionarIntervalo = ttk.Button(self.cal_window, text="Seleccionar Intervalo", command=self.seleccionar_intervalo)
        self.btnSeleccionarIntervalo.pack(pady=10)
        # Asociar la tecla "Enter" al botón
        self.cal_window.bind('<Return>', self.on_enter_key)
        
    def on_enter_key(self, event):
        self.seleccionar_intervalo()

    def seleccionar_intervalo(self):
        fecha_inicio = self.cal_inicio.get_date()
        fecha_fin = self.cal_fin.get_date()
        self.load_records_by_interval(fecha_inicio, fecha_fin)
        self.cal_window.destroy()

    def load_today_records(self):
        today = datetime.now().strftime("%Y-%m-%d")
        self.load_records_by_date(today)

    def load_records_by_date(self, date):
        records = self.db.fetch_records_by_date(date)
        self.update_treeview(records)

    def load_records_by_interval(self, fecha_inicio, fecha_fin):
        records = self.db.fetch_records_by_interval(fecha_inicio, fecha_fin)
        self.update_treeview(records)

    def update_treeview(self, records):
        for record in self.tree.get_children():
            self.tree.delete(record)
        for record in records:
            self.tree.insert("", "end", values=(record[1], record[2], record[3], record[4], record[5], record[6]))

    def convertir_a_pdf(self):
        records = [self.tree.item(item)["values"] for item in self.tree.get_children()]
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        pdf.cell(200, 10, txt="Reporte de DNI", ln=True, align='C')

        headers = ["DNI", "Apellido Paterno", "Apellido Materno", "Nombres", "Fecha y Hora de Ingreso", "Fecha y Hora de Salida"]
        for header in headers:
            pdf.cell(40, 10, header, border=1)
        pdf.ln()

        for record in records:
            for item in record:
                pdf.cell(40, 10, str(item), border=1)
            pdf.ln()

        filename = f"reporte_dni_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        pdf.output(filename)
        messagebox.showinfo("Información", f"El reporte se ha guardado como {filename}")
        
    def convertir_a_excel(self):
        records = [self.tree.item(item)["values"] for item in self.tree.get_children()]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de DNI"

        headers = ["DNI", "Apellido Paterno", "Apellido Materno", "Nombres", "Fecha y Hora de Ingreso", "Fecha y Hora de Salida"]
        sheet.append(headers)

        for record in records:
            sheet.append(record)

        filename = f"reporte_dni_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(filename)
        messagebox.showinfo("Información", f"El reporte se ha guardado como {filename}")

    def ir_a_registro_asistencia(self):
        self.root.destroy()
        subprocess.Popen(["python", "registro_asistencia.py"])

    def __del__(self):
        self.db.close()

if __name__ == "__main__":
    def restart_login():
        root = Tk()
        app = Login(root)
        root.mainloop()

    root = Tk()
    app = FrmReporteDNI(root, restart_login)
    root.mainloop()

    root = Tk()
    app = FrmReporteDNI(root, restart_login)
    root.mainloop()
